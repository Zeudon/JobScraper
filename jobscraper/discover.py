"""Discover careers page URLs for companies missing them."""

import asyncio

from playwright.async_api import async_playwright

from jobscraper.config import get_env

PAGE_LOAD_TIMEOUT = 20_000


async def find_careers_url(company_name: str) -> str | None:
    """Find a company's careers page URL.

    Uses DuckDuckGo (no CAPTCHA) as primary, with LLM fallback.
    """
    # Try DuckDuckGo first
    url = await _search_duckduckgo(company_name)
    if url:
        return url

    # Fallback: ask the LLM (it often knows common careers URLs)
    url = _ask_llm_for_careers_url(company_name)
    return url


async def _search_duckduckgo(company_name: str) -> str | None:
    """Search DuckDuckGo for the company's careers page."""
    headless = get_env("BROWSER_MODE", "headed") != "headed"
    query = f"{company_name} careers page jobs"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        try:
            search_url = f"https://duckduckgo.com/?q={query.replace(' ', '+')}"
            await page.goto(search_url, wait_until="domcontentloaded",
                            timeout=PAGE_LOAD_TIMEOUT)
            await asyncio.sleep(4)

            # Extract search result links
            links = await page.evaluate("""() => {
                const results = document.querySelectorAll('a[href]');
                const urls = [];
                for (const a of results) {
                    const href = a.href;
                    if (href && href.startsWith('http') &&
                        !href.includes('duckduckgo.com') &&
                        !href.includes('google.com') &&
                        !href.includes('youtube.com') &&
                        !href.includes('wikipedia.org') &&
                        !href.includes('glassdoor') &&
                        !href.includes('indeed.com') &&
                        !href.includes('linkedin.com') &&
                        !href.includes('bing.com')) {
                        urls.push(href);
                    }
                }
                return urls.slice(0, 15);
            }""")

            return _pick_best_careers_url(links, company_name)

        except Exception as e:
            print(f"    DuckDuckGo search error: {e}")
            return None
        finally:
            await browser.close()


def _ask_llm_for_careers_url(company_name: str) -> str | None:
    """Ask the LLM for the company's careers page URL as a fallback."""
    from jobscraper.extractor import _call_llm, _parse_json_from_response

    prompt = f"""What is the official careers/jobs page URL for {company_name}?

Return a JSON object:
{{"url": "https://..."}}

Rules:
- Return the OFFICIAL careers page of the company (not LinkedIn, Glassdoor, Indeed, etc.)
- The URL should go directly to their job listings or main careers page
- If you're not sure, return {{"url": null}}
- Return ONLY the JSON object.
"""
    try:
        response = _call_llm(prompt)
        data = _parse_json_from_response(response)
        url = data.get("url")
        if url and url.startswith("http"):
            return url
    except Exception:
        pass
    return None


def _pick_best_careers_url(urls: list[str], company_name: str) -> str | None:
    """Pick the most likely careers page URL from search results."""
    if not urls:
        return None

    company_lower = company_name.lower().replace(" ", "")

    # Priority 1: URL contains "careers" or "jobs" and the company name
    for url in urls:
        url_lower = url.lower()
        if any(kw in url_lower for kw in ["career", "jobs", "job", "hiring", "openings"]):
            # Check if URL seems to belong to the company
            if any(part in url_lower for part in [company_lower, company_lower.replace(".", "")]):
                return url

    # Priority 2: URL contains "careers" or "jobs" (may be the company's site)
    for url in urls:
        url_lower = url.lower()
        if any(kw in url_lower for kw in ["career", "jobs"]):
            return url

    # Priority 3: First result that seems to be the company's domain
    for url in urls:
        url_lower = url.lower()
        if company_lower in url_lower:
            return url

    # Fallback: first result
    return urls[0] if urls else None


async def discover_missing_careers_urls(companies: list) -> list:
    """For companies missing a careers URL, search and fill it in.

    Returns the updated companies list. Also updates the Excel file
    with discovered URLs so they persist for future runs.
    """
    from jobscraper.config import COMPANIES_FILE
    from openpyxl import load_workbook

    missing = [c for c in companies if not c.careers_url]
    if not missing:
        return companies

    print(f"\nDiscovering careers pages for {len(missing)} companies...")

    # Find URLs
    updates = {}  # company_name -> discovered_url
    for company in missing:
        print(f"  Searching: {company.name}...")
        url = await find_careers_url(company.name)
        if url:
            company.careers_url = url
            updates[company.name] = url
            print(f"    Found: {url}")
        else:
            print(f"    Not found — skipping {company.name}")

    # Write discovered URLs back to the Excel file
    if updates:
        try:
            wb = load_workbook(COMPANIES_FILE)
            ws = wb.active

            # Find column indices
            headers = [str(cell.value).strip().lower() if cell.value else ""
                       for cell in ws[1]]
            name_idx = headers.index("company") + 1 if "company" in headers else 1
            url_idx = headers.index("careers url") + 1 if "careers url" in headers else 2

            for row in range(2, ws.max_row + 1):
                name = str(ws.cell(row=row, column=name_idx).value or "").strip()
                if name in updates:
                    ws.cell(row=row, column=url_idx, value=updates[name])

            wb.save(COMPANIES_FILE)
            wb.close()
            print(f"  Updated {len(updates)} URLs in Preferred_Companies.xlsx")
        except Exception as e:
            print(f"  Warning: Could not update Excel file: {e}")

    # Remove companies that still have no URL
    companies = [c for c in companies if c.careers_url]
    return companies
