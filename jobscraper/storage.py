"""Excel storage for job openings — read, deduplicate, and append."""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from jobscraper.config import OPENINGS_FILE

HEADERS = ["Job ID", "Company", "Title", "URL", "Location",
           "Experience", "Date Posted", "Date Scraped"]


@dataclass
class JobOpening:
    """A single job opening."""
    job_id: str  # URL-based or hash-based unique identifier
    company: str
    title: str
    url: str
    location: str = ""
    experience: str = ""
    date_posted: str = ""
    date_scraped: str = ""


def _ensure_file() -> Path:
    """Create the Job_Openings.xlsx file with headers if it doesn't exist."""
    if not OPENINGS_FILE.exists():
        OPENINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Openings"
        ws.append(HEADERS)
        wb.save(OPENINGS_FILE)
        wb.close()
    return OPENINGS_FILE


def load_existing_urls() -> set[str]:
    """Load all existing job URLs from the spreadsheet for dedup."""
    path = _ensure_file()
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    urls = set()
    url_idx = HEADERS.index("URL")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and url_idx < len(row) and row[url_idx]:
            urls.add(str(row[url_idx]).strip())
    wb.close()
    return urls


def save_new_jobs(jobs: list[JobOpening]) -> int:
    """Append new (non-duplicate) jobs to the spreadsheet. Returns count saved."""
    if not jobs:
        return 0

    path = _ensure_file()
    existing_urls = load_existing_urls()

    # Filter out URLs already persisted AND de-duplicate within this batch —
    # the same job can arrive twice (e.g. overlapping "Load More" snapshots or
    # a job listed under multiple department pages), which would otherwise
    # write duplicate rows.
    new_jobs = []
    seen_urls: set[str] = set()
    for job in jobs:
        url = (job.url or "").strip()
        if not url or url in existing_urls or url in seen_urls:
            continue
        seen_urls.add(url)
        new_jobs.append(job)

    if not new_jobs:
        return 0

    wb = load_workbook(path)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")
    for job in new_jobs:
        ws.append([
            job.job_id,
            job.company,
            job.title,
            job.url,
            job.location,
            job.experience,
            job.date_posted or "",
            job.date_scraped or today,
        ])

    wb.save(path)
    wb.close()
    return len(new_jobs)
