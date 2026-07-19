"""Load configuration from YAML filters and Excel companies list."""

import os
from dataclasses import dataclass, field
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

DATA_DIR = Path(__file__).parent.parent / "data"
COMPANIES_FILE = DATA_DIR / "Preferred_Companies.xlsx"
OPENINGS_FILE = DATA_DIR / "Job_Openings.xlsx"
FILTERS_FILE = DATA_DIR / "filters.yaml"


@dataclass
class RolePreference:
    """A single role preference from the YAML config."""
    title: str
    locations: list[str] = field(default_factory=list)
    experience_min: int | None = None
    experience_max: int | None = None
    keywords_include: list[str] = field(default_factory=list)
    keywords_exclude: list[str] = field(default_factory=list)


@dataclass
class Company:
    """A company entry from the Excel file."""
    name: str
    careers_url: str
    roles: list[str] = field(default_factory=list)  # empty = use YAML defaults


def load_filters() -> list[RolePreference]:
    """Load role preferences from filters.yaml."""
    with open(FILTERS_FILE, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    preferences = []
    for pref in data.get("preferences", []):
        exp = pref.get("experience", {})
        keywords = pref.get("keywords", {})
        preferences.append(RolePreference(
            title=pref["title"],
            locations=pref.get("locations", []),
            experience_min=exp.get("min"),
            experience_max=exp.get("max"),
            keywords_include=keywords.get("include", []),
            keywords_exclude=keywords.get("exclude", []),
        ))
    return preferences


def load_companies() -> list[Company]:
    """Load companies from Preferred_Companies.xlsx.

    Expected columns: Company | Careers URL | Roles (optional)
    Roles column can contain comma-separated role titles for per-company filtering.
    """
    wb = load_workbook(COMPANIES_FILE, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    # Find column indices from header row
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_map = {h: i for i, h in enumerate(headers)}

    name_idx = col_map.get("company", 0)
    url_idx = col_map.get("careers url", 1)
    roles_idx = col_map.get("roles", None)

    companies = []
    for row in rows[1:]:
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        url = str(row[url_idx]).strip() if row[url_idx] else ""
        if not name or not url:
            continue

        roles = []
        if roles_idx is not None and roles_idx < len(row) and row[roles_idx]:
            roles = [r.strip() for r in str(row[roles_idx]).split(",") if r.strip()]

        companies.append(Company(name=name, careers_url=url, roles=roles))

    wb.close()
    return companies


def get_env(key: str, default: str = "") -> str:
    return os.getenv(key, default)
