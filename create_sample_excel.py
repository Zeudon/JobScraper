"""One-time script to create a sample Preferred_Companies.xlsx."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Companies"

# Headers
headers = ["Company", "Careers URL", "Roles"]
ws.append(headers)

# Style headers
header_font = Font(bold=True)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Sample companies — edit these with your actual list
# Roles column: comma-separated role restrictions for this company.
#   If empty, all roles from filters.yaml are searched.
#   If filled, only those roles (further restricted by YAML filters) are scraped.
companies = [
    ("Google", "https://www.google.com/about/careers/", "Software Engineer, ML Engineer, AI Engineer"),
    ("Microsoft", "https://careers.microsoft.com/", ""),
    ("Stripe", "https://stripe.com/jobs", "Software Engineer, Forward Deployed Engineer"),
    ("OpenAI", "https://openai.com/careers/", "AI Engineer, Research Engineer"),
    ("Anthropic", "https://www.anthropic.com/careers", "Software Engineer, AI Engineer, Research Engineer"),
    ("Meta", "https://www.metacareers.com/", "Software Engineer, ML Engineer, AI/ML Engineer"),
    ("Apple", "https://jobs.apple.com/", ""),
    ("Amazon", "https://www.amazon.jobs/", "Software Engineer, Applied Scientist"),
    ("Netflix", "https://jobs.netflix.com/", "Software Engineer, Data Scientist"),
    ("Palantir", "https://www.palantir.com/careers/", "Forward Deployed Engineer, Software Engineer"),
]

for name, url, roles in companies:
    ws.append([name, url, roles])

# Auto-size columns
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

wb.save("data/Preferred_Companies.xlsx")
print("Created data/Preferred_Companies.xlsx with", len(companies), "companies")
